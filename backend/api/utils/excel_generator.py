from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from datetime import datetime
import calendar
import os
import time
from pathlib import Path
import platform
from pathlib import Path
from dotenv import load_dotenv

class ExcelHandler:
    def __init__(self):
        self.system = platform.system()

    def save_file(self, wb, filepath):
        """Save Excel file with platform-specific handling"""
        temp_path = f"{filepath}.tmp"
        
        try:
            # Save to temp file first
            wb.save(temp_path)
            
            if self.system == "Windows":
                try:
                    import pythoncom
                    import win32com.client
                    
                    pythoncom.CoInitialize()
                    excel = win32com.client.Dispatch("Excel.Application")
                    for workbook in excel.Workbooks:
                        if workbook.FullName == filepath:
                            workbook.Save()
                    excel.Quit()
                except ImportError:
                    # Fall back to direct file operation
                    os.replace(temp_path, filepath)
            else:
                # Linux handling
                os.replace(temp_path, filepath)
            
            return True
            
        except Exception as e:
            print(f"Save failed: {str(e)}")
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return False
        
class ExcelGenerator:
    load_dotenv()
    BASE_DIR = Path(__file__).resolve().parent.parent.parent.parent
    export_dir = os.path.join(BASE_DIR, 'data')
    OUTPUT_PATH = os.path.join(export_dir, 'excel') + "/"
    DEPT_CODE = "K390140R1C"
    DEPT_NAME = "BG6-RD Center-Automatic System Test R&D Div.1-Dept.1-PTB Sec.1"
    MAX_RETRIES = 3
    RETRY_DELAY = 2  # seconds

    @classmethod
    def is_file_locked(cls, filepath):
        """Check if Excel file is currently open"""
        if not Path(filepath).exists():
            return False
            
        try:
            # Try to open file in write mode
            with open(filepath, 'a+b') as f:
                f.seek(0)  # Try to seek to validate write access
                return False
        except (PermissionError, IOError):
            print(f"File {filepath} is locked or in use")
            return True

    @classmethod
    def generate_excel_files(cls, data, date_obj):
        if not data:
            print(f"No data provided for date {date_obj}, cleaning up files...")
            cls._delete_files(date_obj)
            return None, None
            
        os.makedirs(cls.OUTPUT_PATH, exist_ok=True)
        excel_filename = f"{cls.OUTPUT_PATH}{date_obj.strftime('%Y%m%d')}OT.xlsx"
        summary_filename = f"{cls.OUTPUT_PATH}{date_obj.strftime('%Y%m%d')}OTSummary.xlsx"
        
        excel_handler = ExcelHandler()
        
        for attempt in range(cls.MAX_RETRIES):
            try:
                print(f"Generating Excel files for date {date_obj}...")
                wb_form = cls.create_ot_form(excel_filename, data, date_obj)
                wb_summary = cls.create_ot_summary(summary_filename, data, date_obj)
                
                if not excel_handler.save_file(wb_form, excel_filename):
                    raise PermissionError("Could not save Excel form file")
                if not excel_handler.save_file(wb_summary, summary_filename):
                    raise PermissionError("Could not save Summary file")
                    
                print("Excel files generated successfully")
                return excel_filename, summary_filename
                
            except PermissionError as e:
                print(f"Permission error on attempt {attempt + 1}: {str(e)}")
                if attempt == cls.MAX_RETRIES - 1:
                    raise Exception("Files are in use. Please close them and try again.")
                time.sleep(cls.RETRY_DELAY)
            except Exception as e:
                print(f"Error generating files: {str(e)}")
                raise

    @classmethod
    def _delete_files(cls, date_obj):
        """Delete existing Excel files for given date"""
        try:
            excel_file = f"{cls.OUTPUT_PATH}{date_obj.strftime('%Y%m%d')}OT.xlsx"
            summary_file = f"{cls.OUTPUT_PATH}{date_obj.strftime('%Y%m%d')}OTSummary.xlsx"
            
            for filepath in [excel_file, summary_file]:
                if os.path.exists(filepath) and not cls.is_file_locked(filepath):
                    os.remove(filepath)
                    print(f"Deleted: {filepath}")
        except Exception as e:
            print(f"Error deleting files: {str(e)}")
    
    @classmethod
    def create_ot_form(cls, filename, data, date_info):
        wb = Workbook()
        ws = wb.active
    
        # Column widths setup
        columns = {
            'A': 4.14, 'B': 13.43, 'C': 21.14, 'D': 22.29, 'E': 22.86,
            'F': 10.14, 'G': 15.14, 'H': 13.57, 'I': 15.71, 'J': 1.00,
            'K': 17.14, 'L': 14.43, 'M': 15.43, 'N': 18.57
        }
        for col, width in columns.items():
            ws.column_dimensions[col].width = width

        # Row heights
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 80
        ws.row_dimensions[3].height = 92

        # Title
        ws.merge_cells('A1:N1')
        ws['A1'] = 'Form Lembur (加 班 申 请 单)'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Arial', size=20)

        # Department info
        ws.merge_cells('A2:C2')
        ws['A2'] = f'Departemen\n(部门代码)：\n{cls.DEPT_CODE}\n{cls.DEPT_NAME}'
        ws['A2'].alignment = Alignment(vertical='center', wrap_text=True)
        ws['A2'].font = Font(name='Arial', size=11)

        # Employee classification
        ws['D2'] = 'Klasifikasi Karyawan\n(人员分类):'
        ws['D2'].alignment = Alignment(vertical='center', horizontal='right', wrap_text=True)
        ws['E2'] = '☐ Band0(1-2職等)\n☑ Band1(3-5職等)'
        ws['E2'].alignment = Alignment(vertical='center', wrap_text=True)

        # Date information
        ch_date = f"{date_info.year} 年 {date_info.month:02d} 月 {date_info.day:02d} 日"
        weekday = calendar.day_name[date_info.weekday()].upper()
        ws.merge_cells('F2:H2')
        ws['F2'] = f'Tanggal Lembur （加班日期）:\n                           {ch_date}\nHari （星期）: {weekday}'
        ws['F2'].alignment = Alignment(vertical='center', wrap_text=True)

        print("Excel data received:", data)
        print("First request:", data[0] if data else "No data")

        print("DEBUG - Weekend Check:")
        print(f"Date: {date_info}")
        print(f"Weekday number: {date_info.weekday()}")  # 0=Monday, 5=Saturday, 6=Sunday
        print(f"Is weekend?: {date_info.weekday() >= 5}")
        
        if data and len(data) > 0:
            first_request = data[0]
            print("First request data:", first_request)
            print("is_weekend value:", first_request.get('is_weekend'))
            print("is_holiday value:", first_request.get('is_holiday'))

        # Add overtime type based on is_weekend/is_holiday
        overtime_type = 'Jenis Lembur （加班类别） :\n'
        if data and len(data) > 0:
            first_request = data[0]
            is_weekend = first_request.get('is_weekend', False)
            is_holiday = first_request.get('is_holiday', False)

            if is_holiday:
                overtime_type += ('☐ Saat Hari Kerja (工作日延长加班)\n'
                                '☐ Saat Hari Libur (休息日加班)\n'
                                '☑ Saat Tanggal Merah (法定假日加班)')
            elif is_weekend:
                overtime_type += ('☐ Saat Hari Kerja (工作日延长加班)\n'
                                '☑ Saat Hari Libur (休息日加班)\n'
                                '☐ Saat Tanggal Merah (法定假日加班)')
            else:
                overtime_type += ('☑ Saat Hari Kerja (工作日延长加班)\n'
                                '☐ Saat Hari Libur (休息日加班)\n'
                                '☐ Saat Tanggal Merah (法定假日加班)')
        
        ws.merge_cells('I2:N2')
        ws['I2'] = overtime_type
        ws['I2'].alignment = Alignment(vertical='center', wrap_text=True)

        # Column headers
        headers = [
            ('A3', 'No'),
            ('B3', 'No Karyawan\n(工号)'),
            ('C3', 'Nama\n(姓名)'),
            ('D3', 'Alasan Lembur\n(申请加班事由)'),
            ('E3', 'Jam Lembur (Waktu)\n(预计加班\n起止时间)'),
            ('F3', 'Durasi Lembur\n(预计加班时数)'),
            ('G3', 'Jam istirahat saat lembur (jika perlu, gunakan V)\n(预计休息或用餐打V)'),
            ('H3', 'Tanda Tangan Karyawan\n(员工签名)'),
            ('I3', 'Tanda Tangan Supervisor\n(课级主管签名)'),
            ('K3', 'Konfirmasi Jam Lembur (Waktu)\n(实际加班\n起止时间)'),
            ('L3', 'Konfirmasi Durasi Lembur\n(实际加班时数)'),
            ('M3', 'Jam Istirahat (Jika ada, gunakan V)\n(实际休息或用餐打V)'),
            ('N3', 'Tanda Tangan Karyawan\n(员工签名)')
        ]

        # Apply header formatting
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
        header_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_font = Font(name='Arial', size=11)

        for cell, text in headers:
            ws[cell] = text
            ws[cell].alignment = header_style
            ws[cell].font = header_font
            ws[cell].border = thin_border

        # Add empty rows with formatting
        for row in range(4, 34):
            for col in 'ABCDEFGHIKLMN':
                cell = f'{col}{row}'
                ws[cell] = ''
                ws[cell].border = thin_border
                ws[cell].font = Font(name='Arial', size=11)
                ws[cell].alignment = Alignment(vertical='center')
            ws[f'A{row}'] = row - 3  # Add row numbers

        # Fill in data
        current_row = 4
        for item in data:
            ws[f'B{current_row}'] = item['employee_id']
            ws[f'C{current_row}'] = item['employee_name']
            
            # Add centered cells
            ws[f'D{current_row}'] = item['reason']
            ws[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws[f'E{current_row}'] = f"{item['time_start']} - {item['time_end']}"
            ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws[f'F{current_row}'] = f"{item['total_hours']} hour(s)"
            ws[f'F{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws[f'G{current_row}'] = f"{item['break_start']} - {item['break_end']}" if item['has_break'] else '-'
            ws[f'G{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Add actual time, duration and break status (keep these centered as well)
            ws[f'K{current_row}'] = f"{item['time_start']} - {item['time_end']}"
            ws[f'K{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws[f'L{current_row}'] = f"{item['total_hours']} hour(s)"
            ws[f'L{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws[f'M{current_row}'] = f"{item['break_start']} - {item['break_end']}" if item['has_break'] else '-'
            ws[f'M{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1

        # Footer notes
        ws.merge_cells('A34:N34')
        ws['A34'] = 'Keterangan：'
        ws['A34'].font = Font(name='Arial', size=12)

        notes = [
            ('A35:N35', '1、Informasi diatas harus dilaporkan dengan benar, pelanggaran akan dikenakan sesuai dengan hukuman yang ada dari managemen perusahaan\n(以上资料请据实申报，违者按奖惩管理办法处理)。'),
            ('A36:N36', '2、Karyawan harus menandatangani aplikasi lembur. Jika tidak ada tanda tangan maka akan di anggap tidak sah (实际加班时数，以员工签名确认为准)。'),
            ('A37:N37', '3、Tidak ada aplikasi lembur tidak akan di hitung untuk upah lembur (无加班申请单不计发加班费)。')
        ]

        for cells, text in notes:
            ws.merge_cells(cells)
            ws[cells.split(':')[0]] = text
            ws[cells.split(':')[0]].font = Font(name='Arial', size=12)
            ws[cells.split(':')[0]].alignment = Alignment(vertical='center', wrap_text=True)

        # Form number
        ws['M38'] = 'Form No.:PTB-TB004-001 Rev.01'
        ws['M38'].font = Font(name='Arial', size=10)

        # Save workbook
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        wb.save(filename)

    @classmethod
    def create_ot_summary(cls, filename, data, date_info):
        wb = Workbook()
        ws = wb.active
        
        # Set column widths
        columns = {
            'A': 11.5,  # Work ID
            'B': 15.5,  # Name
            'C': 21.0,  # Project
            'D': 9.75,  # Start Time
            'E': 9.75,  # End Time
            'F': 9.75,  # Break
            'G': 6.0,   # Hours
            'H': 25.0,   # Reason
            'I': 35.0   # Detail
        }
        
        for col, width in columns.items():
            ws.column_dimensions[col].width = width

        # Headers with formatting
        headers = [
            ('A1', 'Work ID'),
            ('B1', 'Name'),
            ('C1', 'Project'),
            ('D1', 'Start Time'),
            ('E1', 'End Time'),
            ('F1', 'Break'),
            ('G1', 'Hours'),
            ('H1', 'Reason'),
            ('I1', 'Detail')
        ]

        header_style = Alignment(horizontal='center', vertical='center')
        header_font = Font(name='Arial', size=11)

        for cell, text in headers:
            ws[cell] = text
            ws[cell].alignment = header_style
            ws[cell].font = header_font

        # Fill data
        current_row = 2
        for item in data:
            ws[f'A{current_row}'] = item['employee_id']
            ws[f'B{current_row}'] = item['employee_name']
            ws[f'C{current_row}'] = item['project']
            ws[f'D{current_row}'] = item['time_start']
            ws[f'E{current_row}'] = item['time_end']
            ws[f'F{current_row}'] = f"{item['break_start']} - {item['break_end']}" if item['has_break'] else '-'
            ws[f'G{current_row}'] = item['total_hours']
            ws[f'H{current_row}'] = item['reason']
            ws[f'I{current_row}'] = item['detail']

            # Formatting
            for col in 'ABCDEFGHI':
                cell = f'{col}{current_row}'
                ws[cell].font = Font(name='Arial', size=11)
                ws[cell].alignment = Alignment(vertical='center')
                if col in 'ADEFG':
                    ws[cell].alignment = Alignment(horizontal='center', vertical='center')

            current_row += 1

        # Save
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        wb.save(filename)

        